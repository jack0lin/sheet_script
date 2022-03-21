# coding:utf-8
import requests
from bs4 import BeautifulSoup
import json
import time, datetime
from datetime import datetime, date, timedelta
import re
import openpyxl
import sys
import tkinter
from tkinter import ttk
from tkinter import messagebox
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import warnings

# 忽略无用警告
warnings.filterwarnings('ignore')


def ck():
    # 根据ck表格cookie创建请求头
    wb_ck = openpyxl.load_workbook('ck_new.xlsx')
    ws_ck = wb_ck['Sheet']
    for row_ck in range(1, ws_ck.max_row + 1):
        cookie = ws_ck.cell(row_ck, 2).value
        cookie = re.findall('com.sankuai.meishi.fe.kdb-bsid=(.*?);', cookie)
        if cookie != []:
            cookie = cookie[0]
            print('---------成功获取cookie---------')
        else:
            print('---------cookie过期,请重新复制---------')
            ws_ck.cell(row_ck, 3).value = str(
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(time.time())))) + '  cookie复制错了,请重新复制'
            sys.exit()
        header_tuiguang = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36',
            'Cookie': 'adpbsid=' + cookie
        }
        header_jingying = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36',
            'Cookie': 'com.sankuai.meishi.fe.kdb-bsid=' + cookie
        }

        # 获取门店各分店名称及poiId
        try:
            url_shopname = 'https://ecom.meituan.com/meishi/gw/rpc/home/-/TEcomOperationDataService/getCityPoiIndex?_tm' + str(
                int(round(time.time() * 1000)))
            rep_shopname = requests.get(url_shopname, headers=header_jingying)
            dict_shopname_poiId = {}
            # print(rep_shopname.text)
            # for city_eachshopname in json.loads(rep_shopname.text.encode("gbk", 'ignore').decode("gbk", "ignore"))['data']:
            # print(json.loads(rep_shopname.text)['data'])
            # time.sleep(5)
            for city_eachshopname in json.loads(rep_shopname.text)['data']:
                for eachshopname in city_eachshopname['pois']:
                    if eachshopname['needCharge'] == False:
                        dict_shopname_poiId[eachshopname['poiName']] = eachshopname['poiId']
        # print('获取poiId列表：'+str(dict_shopname_poiId))
        except Exception as e:
            print(e)
            print('请重新复制cookie')
            continue

        # 获取所有分店推广通launchIds
        url_report = 'https://midas.dianping.com/shopdiy-node/report'
        rep_report = requests.get(url_report, headers=header_tuiguang)
        soup_report = BeautifulSoup(rep_report.text)
        json_report = str(soup_report.find_all('script', id='__NEXT_DATA__')[0])
        json_report = json.loads(re.findall('(\\{.*\\})', json_report)[0])
        dict_shopname_launchId = {}
        for launchList in json_report['props']['initialState']['report']['shopLaunchList']:
            launchList_shopname = re.sub('\\(', '（', launchList['shopName'])
            launchList_shopname = re.sub('\\)', '）', launchList_shopname)
            total_launchId = ''
            for launchId in launchList['launchList']:
                total_launchId = total_launchId + ',' + str(launchId['launchId']) + ','
            total_launchId1 = re.sub(',,', '。。', total_launchId)
            total_launchId2 = re.sub(',', '', total_launchId1)
            total_launchId3 = re.sub('。。', ',', total_launchId2)
            dict_shopname_launchId[launchList_shopname] = total_launchId3
        # print('获取launchId列表：'+str(dict_shopname_launchId))

        # 获取所有合同
        url_contract = 'https://ecom.meituan.com/finance/gw/api/daily/-/finance/common/contractdealpoilist?_tm=' + str(
            int(round(time.time() * 1000))) + '&subBizType=0'
        rep_contract = requests.get(url_contract, headers=header_jingying)
        dict_contract = {}
        for each_contract in json.loads(rep_contract.text)['data']['accountMenuList']:
            dict_contract[each_contract['accountName']] = {}
            dict_contract[each_contract['accountName']]['contractId'] = each_contract['contractId']
            dict_contract[each_contract['accountName']]['Id'] = each_contract['id']
        # print('团购合同列表：'+str(dict_contract))
        dict_contractpoiId = {}
        for each_contractpoiId in json.loads(rep_contract.text)['data']['poiList']:
            name_contractpoiId = re.sub('.*_', '', each_contractpoiId['name'])
            dict_contractpoiId[name_contractpoiId] = each_contractpoiId['id']
        # print('团购合同poiId列表：'+str(dict_contractpoiId))

        # 选择分店和合同
        list_contract = []
        for each_conntract in dict_contract.keys():
            list_contract.append(each_conntract)
        if list_contract == []:
            list_contract.append('（无）')

        for shopname in dict_shopname_poiId.keys():
            for contract in list_contract:
                print('------------------------------------------')
                print(shopname)
                poiId = dict_shopname_poiId[shopname]

                try:
                    # 获取分店推广数据
                    url_tuiguang = 'https://midas.dianping.com/shopdiy/report/datareport/pc/ajax/queryTableV5'
                    param_tuiguang = {
                        'beginDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'endDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'platform': '0',
                        'launchIds': dict_shopname_launchId[shopname],
                        'tabIds': 'T30001,T30020,T30003,T30026,T30006,T30007,T30009,T30027,T30002,T30029',
                        'objectUnit': 'account',
                        'timeUnit': 'day'
                    }
                    rep_tuiguang = requests.get(url_tuiguang, headers=header_tuiguang, params=param_tuiguang)
                    # print(rep_tuiguang.text)
                    json_tuiguang = json.loads(rep_tuiguang.text)
                    # print(json_tuiguang)
                    tuiguang_dateA = str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d"))
                    tuiguang_tuiguangfeiB = json_tuiguang['msg']['total']['T30001']
                    tuiguang_dingdanshuC = json_tuiguang['msg']['total']['T30020']
                    tuiguang_zongdingdanshuD = json_tuiguang['msg']['total']['T30029']
                    tuiguang_dianjishuE = json_tuiguang['msg']['total']['T30003']
                    tuiguang_xingqushuF = json_tuiguang['msg']['total']['T30026']
                    tuiguang_tupianshuG = json_tuiguang['msg']['total']['T30006']
                    tuiguang_pinglunshuH = json_tuiguang['msg']['total']['T30007']
                    tuiguang_tuangoushuI = json_tuiguang['msg']['total']['T30009']
                    tuiguang_liulanshuJ = json_tuiguang['msg']['total']['T30027']
                    tuiguang_baoguangshuK = json_tuiguang['msg']['total']['T30002']
                    if int(tuiguang_zongdingdanshuD):
                        tuiguang_shijichengben = str(round(float(tuiguang_tuiguangfeiB)/float(tuiguang_zongdingdanshuD),2))
                    else:
                        tuiguang_shijichengben = 'Nan'

                    if int(tuiguang_dianjishuE):
                        tuiguang_dianjijunjia = str(round(float(tuiguang_tuiguangfeiB)/float(tuiguang_dianjishuE),2))
                    else:
                        tuiguang_dianjijunjia = 'Nan'

                    print('日期：' + tuiguang_dateA)
                    print('推广费：' + tuiguang_tuiguangfeiB)
                    print('团购订单数：' + tuiguang_dingdanshuC)
                    print('总订单数：' + tuiguang_zongdingdanshuD)
                    print('点击数：' + tuiguang_dianjishuE)
                    print('感兴趣数：' + tuiguang_xingqushuF)
                    print('图片点击数：' + tuiguang_tupianshuG)
                    print('评论点击数：' + tuiguang_pinglunshuH)
                    print('团购点击数：' + tuiguang_tuangoushuI)
                    print('浏览数：' + tuiguang_liulanshuJ)
                    print('曝光数：' + tuiguang_baoguangshuK)
                    print('实际获客成本：' + tuiguang_shijichengben)
                    print('点击均价：' + tuiguang_dianjijunjia)
                    ################################

                    # 获取浏览数据
                    time.sleep(0.5)
                    dict_shopname_shopid = {
                        '曲氏老北京涮肉（夏湾店）': 111948742,
                        '于氏老北京涮肉（前山店）': 132876018,
                        '麗莎花园西餐厅（旧物仓店）': 130826100,
                    }
                    liulan_guanggao = 'Nan'
                    liulan_mendian = 'Nan'
                    liulan_zong = 'Nan'
                    if shopname in dict_shopname_shopid.keys():
                        shopid = dict_shopname_shopid[shopname]
                    else:
                        shopid = dict_shopname_poiId[shopname]

                    # print(dict_shopname_poiId[shopname])
                    url_liulan = 'https://midas.dianping.com/shopdiy/report/datareport/pc/ajax/getBigBoardDataV3'
                    param_liulan = {
                        'beginDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'endDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'shopId': shopid
                    }
                    rep_liulan = requests.get(url_liulan, headers=header_tuiguang, params=param_liulan)
                    json_liulan = json.loads(rep_liulan.text)
                    # print(rep_liulan.text)
                    try:
                        liulan_guanggao = json_liulan['msg']['bigBoardData']['data']['lines'][0]['points'][0]
                        liulan_mendian = json_liulan['msg']['bigBoardData']['data']['lines'][1]['points'][0]
                        liulan_zong = str(int(liulan_guanggao) + int(liulan_mendian))
                        print('总浏览量：' + liulan_zong)
                        print('广告浏览量：' + liulan_guanggao)
                        print('自然浏览量：' + liulan_mendian)
                    except Exception as e:
                        print(e)
                        print('接口报错。。。')

                    # 获取地区数据
                    url_liulan_qvyu = 'https://midas.dianping.com/shopdiy/report/datareport/pc/ajax/getOptionForCompeteDataV2'
                    # param_liulan = {
                    # 	'beginDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                    # 	'endDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                    # 	'shopId': shopid
                    # }
                    rep_liulan_qvyu = requests.get(url_liulan_qvyu, headers=header_tuiguang, params=param_liulan)
                    json_liulan = json.loads(rep_liulan_qvyu.text)
                    # print(rep_liulan.text)
                    try:
                        liulan_city = json_liulan['msg']['area'][0]['id']
                        liulan_district = json_liulan['msg']['area'][1]['id']
                        liulan_region = json_liulan['msg']['area'][2]['id']
                        area = liulan_city + ',' + liulan_district + ',' + liulan_region
                        starLevel = json_liulan['msg']['starLevel'][-1]['id']
                        category = json_liulan['msg']['category'][-1]['id']

                    # print('城市：' + liulan_city)
                    # print('地区：' + liulan_district)
                    # print('地域：' + liulan_region)
                    except Exception as e:
                        print(e)
                        print('接口报错。。。')

                    # 获取排名
                    liulan_paiming_diyi = 'Nan'
                    liulan_paiming_mendian = 'Nan'
                    liulan_paiming_junzhi = 'Nan'
                    url_liulan_paiming = 'https://midas.dianping.com/shopdiy/report/datareport/pc/ajax/getCompeteDataV2'
                    param_liulan_paiming = {
                        'beginDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'endDate': str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")),
                        'shopId': shopid,
                        'area': area,
                        'starLevel': starLevel,
                        'category': category,
                    }
                    rep_liulan_paiming = requests.get(url_liulan_paiming, headers=header_tuiguang,
                                                      params=param_liulan_paiming)
                    json_liulan_paiming = json.loads(rep_liulan_paiming.text)
                    # print(rep_liulan.text)
                    try:
                        liulan_paiming_diyi = json_liulan_paiming['msg']['competeData'][0]['value']
                        liulan_paiming_mendian = json_liulan_paiming['msg']['competeData'][1]['value']
                        liulan_paiming_junzhi = json_liulan_paiming['msg']['competeData'][2]['value']

                        print('同行第一：' + liulan_paiming_diyi)
                        print('我的门店：' + liulan_paiming_mendian)
                        print('同行均值：' + liulan_paiming_junzhi)
                    except Exception as e:
                        print(e)
                        print('接口报错。。。')

                    # 获取评分
                    time.sleep(0.5)
                    'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/getFeedbackSummary?_tm=1647760548474'
                    url_dianpingpingjia_details = 'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/getFeedbackSummary'
                    param_dianpingpingjia = {'_tm': str(int(round(time.time() * 1000)))}
                    data_dianpingpingjia_details = {
                        'platform': '1',
                        # 'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '1',
                            'poiId': poiId,
                            # 'referTag': '0',
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_dianpingpingjia_details = requests.post(url_dianpingpingjia_details, headers=header_jingying,
                                                                params=param_dianpingpingjia,
                                                                json=data_dianpingpingjia_details)
                    # print(rep_dianpingpingjia.text)
                    dianpingzongpingjia_kouwei = json.loads(rep_dianpingpingjia_details.text)['data']['subScores'][0][
                        'score']
                    dianpingzongpingjia_huanjing = json.loads(rep_dianpingpingjia_details.text)['data']['subScores'][1][
                        'score']
                    dianpingzongpingjia_fuwu = json.loads(rep_dianpingpingjia_details.text)['data']['subScores'][2][
                        'score']
                    print('口味评分：' + str(dianpingzongpingjia_kouwei))
                    print('环境评分：' + str(dianpingzongpingjia_huanjing))
                    print('服务评分：' + str(dianpingzongpingjia_fuwu))

                    # 获取实收金额
                    time.sleep(0.5)
                    url_shishoujine = 'https://ecom.meituan.com/emis/gw/rpc/TEcomOperationDataService/getEcomConsumeTimeDataRegion'
                    param_shishoujine = {'_tm': str(int(round(time.time() * 1000)))}
                    data_shishoujine = {'optionType': '1', 'poiId': poiId}
                    rep_shishoujine = requests.post(url_shishoujine, headers=header_jingying, params=param_shishoujine,
                                                    json=data_shishoujine)
                    # print(rep_shishoujine.text)
                    shishoujineQ = json.loads(rep_shishoujine.text)['data']['cards'][0]['extraDatas'][0]['value']
                    shishoujineQ = re.sub('¥ ', '', shishoujineQ)
                    print('实收金额：' + shishoujineQ)

                    # 获取收单金额
                    time.sleep(0.5)
                    shoudanjineR = ''
                    url_shoudanjine = 'https://ecom.meituan.com/emis/gw/rpc/TEcomOperationDataService/getEcomConsumeDistrubution'
                    param_shoudanjine = {'_tm': str(int(round(time.time() * 1000)))}
                    data_shoudanjine = {'optionType': '1', 'poiId': poiId}
                    rep_shoudanjine = requests.post(url_shoudanjine, headers=header_jingying, params=param_shoudanjine,
                                                    json=data_shoudanjine)
                    # print(rep_shoudanjine.text)
                    shoudanjineR = ''
                    tuangoujineRS = ''
                    for find_shoudan in json.loads(rep_shoudanjine.text)['data']['distributions']:
                        if find_shoudan['name'] == '美团收单':
                            shoudanjineR = find_shoudan['value']
                            print('收单金额：' + shoudanjineR)
                        if find_shoudan['name'] == '团购':
                            tuangoujineRS = find_shoudan['value']
                            print('团购金额：' + tuangoujineRS)

                    # 获取大众点评评价数
                    # 总评
                    time.sleep(0.5)
                    url_dianpingpingjia = 'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/queryFeedback'
                    param_dianpingpingjia = {'_tm': str(int(round(time.time() * 1000)))}
                    data_dianpingpingjia = {
                        'platform': '1',
                        'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '1',
                            'poiId': poiId,
                            'referTag': '0',
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_dianpingpingjia = requests.post(url_dianpingpingjia, headers=header_jingying,
                                                        params=param_dianpingpingjia, json=data_dianpingpingjia)
                    # print(rep_dianpingpingjia.text)
                    dianpingzongpingjiaW = json.loads(rep_dianpingpingjia.text)['data']['total']
                    print('点评总评价数：' + str(dianpingzongpingjiaW))

                    # 坏评
                    time.sleep(0.5)
                    data_dianpingbadpingjia = {
                        'platform': '1',
                        'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '1',
                            'poiId': poiId,
                            'referTag': '0',
                            'starTag': '0',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_dianpingbadpingjia = requests.post(url_dianpingpingjia, headers=header_jingying,
                                                           params=param_dianpingpingjia, json=data_dianpingbadpingjia)
                    # print(rep_dianpingpingjia.text)
                    dianpingbadpingjiaX = json.loads(rep_dianpingbadpingjia.text)['data']['total']
                    print('点评坏评价数：' + str(dianpingbadpingjiaX))

                    # 精选
                    time.sleep(0.5)
                    data_dianpingjingxuan = {
                        'platform': '1',
                        'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '1',
                            'poiId': poiId,
                            'referTag': '2',
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_dianpingjingxuan = requests.post(url_dianpingpingjia, headers=header_jingying,
                                                         params=param_dianpingpingjia, json=data_dianpingjingxuan)
                    # print(rep_dianpingpingjia.text)
                    dianpingjingxuanY = json.loads(rep_dianpingjingxuan.text)['data']['total']
                    print('点评精选数：' + str(dianpingjingxuanY))

                    # 获取点评星级
                    time.sleep(0.5)
                    url_dianpingxingji = 'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/getFeedbackSummary'
                    param_dianpingxingji = {'_tm': str(int(round(time.time() * 1000)))}
                    data_dianpingxingji = {
                        'platform': '1',
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '1',
                            'poiId': poiId,
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_dianpingxingji = requests.post(url_dianpingxingji, headers=header_jingying,
                                                       params=param_dianpingxingji, json=data_dianpingxingji)
                    # print(rep_meituanxingji.text)
                    dianpingxingjiAC = json.loads(rep_dianpingxingji.text)['data']['avgScore']
                    print('点评星级：' + str(dianpingxingjiAC))

                    # 获取美团评价数
                    # 总评
                    time.sleep(0.5)
                    url_meituanpingjia = 'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/queryFeedback'
                    param_meituanpingjia = {'_tm': str(int(round(time.time() * 1000)))}
                    data_meituanpingjia = {
                        'platform': '0',
                        'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '0',
                            'poiId': poiId,
                            'referTag': '0',
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_meituanpingjia = requests.post(url_meituanpingjia, headers=header_jingying,
                                                       params=param_meituanpingjia, json=data_meituanpingjia)
                    # print(rep_meituanpingjia.text)
                    meituanzongpingjiaAA = json.loads(rep_meituanpingjia.text)['data']['total']
                    print('美团总评价数：' + str(meituanzongpingjiaAA))

                    # 坏评
                    time.sleep(0.5)
                    data_meituanbadpingjia = {
                        'platform': '0',
                        'pageInfo': {'total': '0', 'offset': '0', 'limit': '10'},
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '0',
                            'poiId': poiId,
                            'referTag': '0',
                            'starTag': '0',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_meituanbadpingjia = requests.post(url_meituanpingjia, headers=header_jingying,
                                                          params=param_meituanpingjia, json=data_meituanbadpingjia)
                    # print(rep_meituanbadpingjia.text)
                    meituanbadpingjiaAB = json.loads(rep_meituanbadpingjia.text)['data']['total']
                    print('美团坏评价数：' + str(meituanbadpingjiaAB))

                    # 获取美团星级
                    time.sleep(0.5)
                    url_meituanxingji = 'https://ecom.meituan.com/emis/gw/rpc/TFeedbackEcomService/getFeedbackSummary'
                    param_meituanxingji = {'_tm': str(int(round(time.time() * 1000)))}
                    data_meituanxingji = {
                        'platform': '0',
                        'queryPara': {
                            'businessTag': '-1',
                            'platform': '0',
                            'poiId': poiId,
                            'starTag': '-1',
                            'startTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) * 1000 - 1)
                        }
                    }
                    rep_meituanxingji = requests.post(url_meituanxingji, headers=header_jingying,
                                                      params=param_meituanxingji, json=data_meituanxingji)
                    # print(rep_meituanxingji.text)
                    meituanxingjiAC = json.loads(rep_meituanxingji.text)['data']['avgScore']
                    print('美团星级：' + str(meituanxingjiAC))

                    # 获取排名
                    paimingAE = ''
                    for page_paiming in ['0', '10', '20', '30', '40', '50', '60', '70', '80', '90']:
                        time.sleep(0.1)
                        # 获取排名
                        header_paiming = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36 MicroMessenger/7.0.9.501 NetType/WIFI MiniProgramEnv/Windows WindowsWechat'}
                        url_paimingdianping = 'https://m.dianping.com/wx-business/ranklist/getshoprankinfo'
                        data_paimingdianping = {
                            "rankKey": "a5bee2d22787c997d4a3a1447cdd7b89af9c8673f64b2b89f67a2e713fd8d02f",
                            "start": page_paiming,
                            "pageSize": "10",
                            "come_from": "weixin",
                            "optimus_risk_level": "71",
                            "optimus_code": "10"
                        }
                        rep_paimingdianping = requests.post(url_paimingdianping, headers=header_paiming,
                                                            json=data_paimingdianping)
                        # print(rep_paimingdianping.text)
                        for each_paimingdianping in json.loads(rep_paimingdianping.text)['data']['shopList']:
                            if each_paimingdianping['shopName'] == shopname:
                                # print(each_paimingdianping['shopName'])
                                # print('点评人气榜排名：'+str(each_paimingdianping['rankings']))
                                paimingAE = each_paimingdianping['rankings']
                                paimingname = each_paimingdianping['shopName']
                                break
                        if paimingAE != '':
                            break
                    if paimingAE == '':
                        paimingAE = '100名以外'
                    print('排名：' + str(paimingAE))

                    # 获取套餐价格
                    for each_contractpoiId in dict_contractpoiId.keys():
                        if shopname in each_contractpoiId:
                            poiId_shopname = each_contractpoiId
                            break
                    time.sleep(0.5)
                    if contract != '（无）':
                        # 获取套餐价格
                        url_summarydeallist = 'https://ecom.meituan.com/finance/gw/api/daily/-/finance/profit/summarydeallist'
                        param_summarydeallist = {
                            'accountId': str(dict_contract[contract]['Id']),
                            'contractId': dict_contract[contract]['contractId'],
                            'dealId': '',
                            'dealPartner': '0',
                            'hideZero': '0',
                            'limit': '30',
                            'offset': '0',
                            'partner': '0',
                            'poiId': dict_contractpoiId[poiId_shopname],
                            'search': 'first',
                            'beginTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S")))),
                            'endTime': str(int(time.mktime(time.strptime(
                                str((date.today() + timedelta(days=0)).strftime("%Y-%m-%d")) + ' 00:00:00',
                                "%Y-%m-%d %H:%M:%S"))) - 1)
                        }
                        rep_summarydeallist = requests.get(url_summarydeallist, headers=header_jingying,
                                                           params=param_summarydeallist)
                        # print(rep_summarydeallist.text)
                        dict_tuangou = {}
                        for each_tuangou in json.loads(rep_summarydeallist.text)['data']['list']['items']:
                            if ('已下线' in each_tuangou['name']) == False:
                                each_tuangouming = each_tuangou['name']
                                each_tuangoushu = each_tuangou['details'][0]['cnt']
                                each_tuangoue = each_tuangou['details'][0]['realPayed']
                                each_nameid = each_tuangou['id']
                                dict_tuangou[each_tuangouming] = {}
                                dict_tuangou[each_tuangouming]['cnt'] = each_tuangoushu
                                dict_tuangou[each_tuangouming]['realPayed'] = each_tuangoue
                                dict_tuangou[each_tuangouming]['nameid'] = each_tuangou['id']  # 获取财务管理nameid
                        # print(dict_tuangou)

                        # 获取按团购分析的nameid
                        dict_tuangounameid = {}
                        url_tuangounameid = 'https://ecom.meituan.com/emis/gw/rpc/TEcomBusinessAnalysisService/getAllObjects?_tm=' + str(
                            int(round(time.time() * 1000)))
                        rep_tuangounameid = requests.get(url_tuangounameid, headers=header_jingying)
                        for each_tuangounameid in json.loads(rep_tuangounameid.text)['data']:
                            dict_tuangounameid[each_tuangounameid['bObjectId']] = each_tuangounameid['name']
                        # print(dict_tuangounameid)

                        # 改名
                        dict_changenametuangou = {}
                        list_changenametuangou = []
                        for namefromcaiwu in dict_tuangou.keys():
                            for namefromtuangou in dict_tuangounameid.keys():
                                if dict_tuangou[namefromcaiwu]['nameid'] == namefromtuangou:
                                    print('团购名：' + dict_tuangounameid[namefromtuangou])
                                    print('团购数：' + str(dict_tuangou[namefromcaiwu]['cnt']))
                                    print('团购总金额：' + str(dict_tuangou[namefromcaiwu]['realPayed']))
                                    if dict_tuangounameid[namefromtuangou] in list_changenametuangou:
                                        # print(namefromtuangou)
                                        # print(dict_tuangounameid[namefromtuangou])
                                        # print(dict_tuangou[namefromcaiwu]['cnt'])
                                        # print(dict_tuangou[namefromcaiwu]['realPayed'])
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['cnt'] = \
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['cnt'] + \
                                        dict_tuangou[namefromcaiwu]['cnt']
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['realPayed'] = \
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['realPayed'] + \
                                        dict_tuangou[namefromcaiwu]['realPayed']
                                    else:
                                        # print(namefromcaiwu+' 改为： '+namefromtuangou)
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]] = {}
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['cnt'] = \
                                        dict_tuangou[namefromcaiwu]['cnt']
                                        dict_changenametuangou[dict_tuangounameid[namefromtuangou]]['realPayed'] = \
                                        dict_tuangou[namefromcaiwu]['realPayed']
                                    list_changenametuangou.append(dict_tuangounameid[namefromtuangou])
                    # print(dict_changenametuangou)

                    else:
                        # 获取套餐及代金券
                        url_taocan = 'https://ecom.meituan.com/emis/gw/rpc/TEcomBusinessAnalysisService/getTPageGroupCompareData'
                        param_taoacan = {'_tm': str(int(round(time.time() * 1000)))}
                        data_taocan = {
                            'pageDomain': {'currPage': '0', 'pageSize': '30', 'totalNum': '0'},
                            'sortKey': 'groupVisitUv',
                            'sortOrder': 'desc',
                            'tOptionType': '1'
                        }
                        rep_taocan = requests.post(url_taocan, headers=header_jingying, params=param_taoacan,
                                                   json=data_taocan)
                        # print(rep_taocan.text)
                        dict_changenametuangou = {}
                        for each_taocan in json.loads(rep_taocan.text)['data']['tGroupCompareDatas']:
                            print('套餐名称：' + each_taocan['groupName'])
                            print('套餐售卖：' + str(each_taocan['consuCpnCnt']))
                            print('套餐售卖金额：' + str(each_taocan['consuAmt']))
                            dict_changenametuangou[each_taocan['groupName']] = {}
                            dict_changenametuangou[each_taocan['groupName']]['cnt'] = str(each_taocan['consuCpnCnt'])
                            dict_changenametuangou[each_taocan['groupName']]['realPayed'] = str(each_taocan['consuAmt'])
                    # print(dict_changenametuangou)
                except Exception as e:
                    print('---------后台无相应数据，跳过---------')
                    print(e)
                    continue

                # 导出数据
                try:
                    wb = openpyxl.load_workbook(ws_ck.cell(row_ck, 1).value + '.xlsx')
                    if shopname in wb.get_sheet_names():
                        ws = wb[shopname]
                    else:
                        ws = wb.create_sheet(shopname)
                        ws.append(['日期', '推广通费用（元）', '团购订单数（个）', '总订单数（个）', '广告点击数量（元/次）', '感兴趣数量（次）', '图片点击数量（次）',
                                   '评论点击数量（次）', '团购点击数量（次）', '浏览数量（次）', '曝光数量（次）', '实际获客成本（元）', '点击均价', '总浏览量', '广告浏览量',
                                   '自然浏览量', '总实收金额', '套餐实收金额', '代金券实收金额', '美团收单金额', '同行第一', '我的门店', '同行均值', '新增评价数（点评）',
                                   '新增差评数（点评）', '新增精选评价数（点评）', '星级评分（点评）', '口味评分', '环境评分', '服务评分', '新增评价数量（美团）',
                                   '新增差评数（美团）', '星级评分（美团）', '线上活动', '推广通调整动作', '人气榜（美团）', '点评品类热门榜',
                                   '点评全美食热门榜（全城不分品种总排位）'])
                        try:
                            wb.save(ws_ck.cell(row_ck, 1).value + '.xlsx')
                        except:
                            print(title='提示', message='请关闭表格重新保存')
                        wb = openpyxl.load_workbook(ws_ck.cell(row_ck, 1).value + '.xlsx')
                        ws = wb[shopname]
                except:
                    wb = openpyxl.Workbook(ws_ck.cell(row_ck, 1).value + '.xlsx')
                    ws = wb.create_sheet(shopname)
                    ws.append(
                        ['日期', '推广通费用（元）', '团购订单数（个）', '总订单数（个）', '广告点击数量（元/次）', '感兴趣数量（次）', '图片点击数量（次）', '评论点击数量（次）',
                         '团购点击数量（次）', '浏览数量（次）', '曝光数量（次）', '实际获客成本（元）', '点击均价', '总浏览量', '广告浏览量', '自然浏览量', '总实收金额',
                         '套餐实收金额', '代金券实收金额', '美团收单金额', '同行第一', '我的门店', '同行均值', '新增评价数（点评）', '新增差评数（点评）', '新增精选评价数（点评）',
                         '星级评分（点评）', '口味评分', '环境评分', '服务评分', '新增评价数量（美团）', '新增差评数（美团）', '星级评分（美团）', '线上活动', '推广通调整动作',
                         '人气榜（美团）', '点评品类热门榜', '点评全美食热门榜（全城不分品种总排位）'])

                    try:
                        wb.save(ws_ck.cell(row_ck, 1).value + '.xlsx')
                    except:
                        print(title='提示', message='请关闭表格重新保存')
                    wb = openpyxl.load_workbook(ws_ck.cell(row_ck, 1).value + '.xlsx')
                    ws = wb[shopname]
                row_save = ws.max_row + 1
                ws.cell(row_save, 1).value = str(tuiguang_dateA)
                ws.cell(row_save, 2).value = str(tuiguang_tuiguangfeiB)
                ws.cell(row_save, 3).value = str(tuiguang_dingdanshuC)
                ws.cell(row_save, 4).value = str(tuiguang_zongdingdanshuD)
                ws.cell(row_save, 5).value = str(tuiguang_dianjishuE)
                ws.cell(row_save, 6).value = str(tuiguang_xingqushuF)
                ws.cell(row_save, 7).value = str(tuiguang_tupianshuG)
                ws.cell(row_save, 8).value = str(tuiguang_pinglunshuH)
                ws.cell(row_save, 9).value = str(tuiguang_tuangoushuI)
                ws.cell(row_save, 10).value = str(tuiguang_liulanshuJ)
                ws.cell(row_save, 11).value = str(tuiguang_baoguangshuK)
                ws.cell(row_save, 12).value = str(tuiguang_shijichengben)
                ws.cell(row_save, 13).value = str(tuiguang_dianjijunjia)
                ws.cell(row_save, 14).value = str(liulan_zong)
                ws.cell(row_save, 15).value = str(liulan_guanggao)
                ws.cell(row_save, 16).value = str(liulan_mendian)
                ws.cell(row_save, 17).value = str(shishoujineQ)
                ws.cell(row_save, 18).value = str(tuangoujineRS)
                ws.cell(row_save, 19).value = str(tuangoujineRS)
                ws.cell(row_save, 20).value = str(shoudanjineR)
                ws.cell(row_save, 21).value = str(liulan_paiming_diyi)
                ws.cell(row_save, 22).value = str(liulan_paiming_mendian)
                ws.cell(row_save, 23).value = str(liulan_paiming_junzhi)
                ws.cell(row_save, 24).value = str(dianpingzongpingjiaW)
                ws.cell(row_save, 25).value = str(dianpingbadpingjiaX)
                ws.cell(row_save, 26).value = str(dianpingjingxuanY)
                ws.cell(row_save, 27).value = str(dianpingxingjiAC)
                ws.cell(row_save, 28).value = str(dianpingzongpingjia_kouwei)
                ws.cell(row_save, 29).value = str(dianpingzongpingjia_huanjing)
                ws.cell(row_save, 30).value = str(dianpingzongpingjia_fuwu)
                ws.cell(row_save, 31).value = str(meituanzongpingjiaAA)
                ws.cell(row_save, 32).value = str(meituanbadpingjiaAB)
                ws.cell(row_save, 33).value = str(meituanxingjiAC)
                ws.cell(row_save, 38).value = str(paimingAE)
                columnFORtuangou = 0
                list_columnFORtuangou = []
                while True:
                    if ws.cell(1, 39 + columnFORtuangou).value != None:
                        list_columnFORtuangou.append(columnFORtuangou)
                        columnFORtuangou = columnFORtuangou + 3
                    elif ws.cell(1, 39 + columnFORtuangou).value == None:
                        break

                for columnFORtuangou in list_columnFORtuangou:
                    for each_tuangouSAVE in dict_changenametuangou.keys():
                        if ws.cell(1, 39 + columnFORtuangou).value == each_tuangouSAVE:
                            ws.cell(row_save, 39 + columnFORtuangou).value = str(each_tuangouSAVE)
                            ws.cell(row_save, 40 + columnFORtuangou).value = str(
                                dict_changenametuangou[each_tuangouSAVE]['cnt'])
                            ws.cell(row_save, 41 + columnFORtuangou).value = str(
                                dict_changenametuangou[each_tuangouSAVE]['realPayed'])
                            columnFORtuangou = columnFORtuangou + 3
                try:
                    wb.save(ws_ck.cell(row_ck, 1).value + '.xlsx')
                    print('保存成功')
                except:
                    print('请关闭表格重新保存')


def mail():
    mail_host = "smtp.126.com"  # 设置服务器
    mail_user = "xhtdtk@126.com"  # 用户名
    mail_pass = "BKFJMJRQRNRHAPTH"  # 授权码而非邮箱密码
    receivers = ['814742494@qq.com',
                 '739889363@qq.com']  # -----------------------------------------------------------------------------------------邮箱'814742494@qq.com

    msg = MIMEMultipart()  # 创建一个带附件的实例
    msg["Subject"] = str((date.today() + timedelta(days=-1)).strftime("%Y-%m-%d")) + "（昨日）美团后台记录数据"  # 指定邮件主题
    msg["From"] = mail_user  # 邮件发送人
    msg["To"] = ','.join(receivers)  # 邮件接收人，如果存在多个收件人，可用join连接

    # ---文字部分---
    part = MIMEText("请查收，谢谢！")
    msg.attach(part)

    # ---附件部分---
    for each_fujian in ['ck_new.xlsx', 'Johy咖啡_新.xlsx', '凡熙餐厅_新.xlsx', '啡与啡_新.xlsx', '海王新粤排档_新.xlsx', '火凤祥_新.xlsx', '老北京涮肉_新.xlsx',
                        '叻三度_新.xlsx', '彭渔宴_新.xlsx', '烧江南烤肉_新.xlsx', '泰和楼私房菜_新.xlsx', '五洲鼎食_新.xlsx', '想见你一面_新.xlsx', '丽莎花园西餐_新.xlsx',
                        '寻乐事务所_新.xlsx',
                        '茗粤海鲜酒楼_新.xlsx']:  # ------------------------------------------------------------------------文件
        part = MIMEApplication(open(each_fujian, 'rb').read())
        part.add_header('Content-Disposition', 'attachment', filename=each_fujian)
        msg.attach(part)

    try:
        s = smtplib.SMTP_SSL("smtp.126.com", 465)  # 连接smtp邮件服务器,端口默认是25
        s.login(mail_user, mail_pass)  # 登陆服务器
        s.sendmail(mail_user, receivers, msg.as_string())  # 发送邮件
        s.close()
        print('---------发送邮件成功---------')
    except Exception as e:
        print('---------发送邮件失败---------')
        print(e)


print('Here we go.')
while True:
    if time.strftime('%H', time.localtime(time.time())) == '11':  # --------------------------------------------发送邮箱时间
        ck()
        time.sleep(2)
        mail()
        for numtime in range(70):
            time.sleep(60)
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + '  获取数据后休息70分钟')
    else:
        time.sleep(60)
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())) + '  等待15点到来更新数据')
