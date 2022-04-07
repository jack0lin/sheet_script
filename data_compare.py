# coding:utf-8

import time, datetime
from datetime import datetime, date, timedelta
from openpyxl.styles import Font
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import warnings
from decimal import *

# 忽略无用警告


warnings.filterwarnings('ignore')


def get_template(filename, sheetname, startdate1, enddate1, startdate2, enddate2):

    wb = openpyxl.Workbook(sheetname + startdate1 + 'to' + enddate1 + 'with' + startdate2 + 'to' + enddate2 + '.xlsx')
    ws = wb.create_sheet(sheetname + startdate1 + 'to' + enddate1)
    ws.row_dimensions[1].height = 40
    Color = ['FFFF00', 'F2DCDB']
    Color_row_column = [{
        'color': 'FFFF00',
        'cells': {
            'row': [1, 1+1],
            'column': [1, 9 + 1]
        }
    },
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [2, 2+1],
                'column': [1, 9 + 1]
            },
            # {
            #     'row': [2],
            #     'column': [1, 8]
            # },
        },
    ]
    cell_merge = ["A1:G1", "H1:I1", "A2:I2"]

    fix_data = [[sheetname + '总结（' + startdate1 + '到' + enddate1 + '）','', '', '', '', '', '',
                 '提升/下降\n标橙/标蓝'],
                ['一、浏览量数据', ],
                ['日期', '曝光人数', '点击', '订单量', '访问转化率', '消费转化率', '推广通花费', '商户浏览量', '点击均价',],

                ]
    for row in fix_data:
        ws.append(row)

    wb.save(sheetname + startdate1 + 'to' + enddate1 + 'with' + startdate2 + 'to' + enddate2 + '.xlsx')

    wb = openpyxl.load_workbook(
        filename=sheetname + startdate1 + 'to' + enddate1 + 'with' + startdate2 + 'to' + enddate2 + '.xlsx')

    # 获取所有的sheet名
    sheets = wb.sheetnames
    # 读取第一个sheet表格
    ws = wb[sheets[0]]

    font_pass = Font(color="FF2600")

    rows1, rows11, rows111, rows1111, rows11111 = get_data(filename, sheetname, startdate1, enddate1)
    rows2, rows22, rows222, rows2222, rows22222 = get_data(filename, sheetname, startdate2, enddate2)
    for row in rows1:
        ws.append(row)
    ws.append(get_rows_sum(rows1))
    ws.append(get_rows_sum(rows2))
    fix_data2 = ['总结', '曝光人数环比', '点击环比', '订单量环比', '访问转化率环比', '消费转化率环比', '推广通花费环比', '商户浏览量环比', '', ]
    ws.append(fix_data2)
    row_ratio = ['',
                 get_ratio(get_rows_sum(rows1)[1], get_rows_sum(rows2)[1]),
                 get_ratio(get_rows_sum(rows1)[2], get_rows_sum(rows2)[2]),
                 get_ratio(get_rows_sum(rows1)[3], get_rows_sum(rows2)[3]),
                 get_ratio(float(get_rows_sum(rows1)[4].replace("%", "")), float(get_rows_sum(rows2)[4].replace("%", ""))),
                 get_ratio(float(get_rows_sum(rows1)[5].replace("%", "")), float(get_rows_sum(rows2)[5].replace("%", ""))),
                 get_ratio(get_rows_sum(rows1)[6], get_rows_sum(rows2)[6]),
                 get_ratio(get_rows_sum(rows1)[7], get_rows_sum(rows2)[7]),
                 ]
    ws.append(row_ratio)

    cell_merge.append("A" + str(len(rows1)+7-1) + ':A' + str(len(rows1)+7+4))
    cell_merge.append("B" + str(len(rows1)+7+1) + ':I' + str(len(rows1)+7+4))


    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 6, len(rows1) + 12],
                'column': [1, 9 + 1]
            }
        }
    )

    summary1 = "总结：近" + str(len(rows1)) + "天（" + startdate1 + '到' + enddate1 + '）较上一轮'\
               + str(len(rows2)) + "天（" + startdate2 + '到' + enddate2 + '）'\
               + fix_data2[1] + get_change_text(row_ratio[1]) + ','\
               + fix_data2[2] + get_change_text(row_ratio[2]) + ','\
               + fix_data2[3] + get_change_text(row_ratio[3]) + ','\
               + fix_data2[4] + get_change_text(row_ratio[4]) + ','\
               + fix_data2[5] + get_change_text(row_ratio[5]) + ','\
               + fix_data2[6] + get_change_text(row_ratio[6]) + ','\
               + fix_data2[7] + get_change_text(row_ratio[7]) + '.'

    ws.append(['', summary1])
    ws.cell(row=len(rows1) + 8, column=2).font = font_pass

    fix_data3 = [
                [],
                [],
                [],
                [],
                ['二、门店浏览量', ],
                ['', '浏览量', '广告带来浏览量', '自然流量', '日均浏览量', '广告占比', '同行第一', '', '', ],
                ]
    for row in fix_data3:
        ws.append(row)

    ws.append(get_rows_sum2(rows11))
    ws.append(get_rows_sum2(rows22))
    fix_data4 = ['总结', '浏览量环比', '广告带来浏览量环比', '自然流量环比',]
    ws.append(fix_data4)
    row_ratio = ['',
                 get_ratio(get_rows_sum2(rows11)[1], get_rows_sum2(rows22)[1]),
                 get_ratio(get_rows_sum2(rows11)[2], get_rows_sum2(rows22)[2]),
                 get_ratio(get_rows_sum2(rows11)[3], get_rows_sum2(rows22)[3]),
                 ]
    ws.append(row_ratio)
    summary2 = "总结：近" + str(len(rows11)) + "天（" + startdate1 + '到' + enddate1 + '）较上一轮' \
               + str(len(rows22)) + "天（" + startdate2 + '到' + enddate2 + '）' \
               + fix_data3[5][1] + get_change_text(row_ratio[1]) + ',' \
               + fix_data3[5][2] + get_change_text(row_ratio[2]) + ',' \
               + fix_data3[5][3] + get_change_text(row_ratio[3]) + ',' \
               + fix_data3[5][4] + str(get_rows_sum2(rows11)[4]) + ',' \
               + fix_data3[5][5] + str(get_rows_sum2(rows11)[5]) + '.'

    ws.append(['', summary2])
    ws.cell(row=len(rows1) + 7 + 4 + 3 + 5, column=2).font = font_pass

    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2) + ':I' + str(len(rows1) + 7 + 4 + 2))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 6) + ':A' + str(len(rows1) + 7 + 4 + 5 + 5))
    cell_merge.append("B" + str(len(rows1) + 7 + 4 + 8) + ':I' + str(len(rows1) + 7 + 4 + 10))
    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2, len(rows1) + 7 + 4 + 2 + 1],
                'column': [1, 9 + 1]
            }
        }
    )
    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 6, len(rows1) + 7 + 4 + 2 + 8],
                'column': [1, 9 + 1]
            }
        }
    )

    fix_data5 = [
        [],
        [],
        [],
        ['三、热门榜/评价榜排名分析', ],
        ['日期', '类目', '热门榜',  ],
    ]
    for row in fix_data5:
        ws.append(row)
    ws.append([rows111[-1][0], rows111[-1][2], '第' + rows111[-1][1] + '名',])
    summary3 = ["总结：" + rows111[-1][2] + fix_data5[4][1] + "，榜单展示：" + fix_data5[4][2] + '第' + rows111[-1][1] + '名']
    ws.append(summary3)
    ws.cell(row=len(rows1) + 7 + 4 + 3 + 5 + 7, column=1).font = font_pass

    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10))
    cell_merge.append("C" + str(len(rows1) + 7 + 4 + 2 + 11) + ':I' + str(len(rows1) + 7 + 4 + 2 + 11))
    cell_merge.append("C" + str(len(rows1) + 7 + 4 + 2 + 12) + ':I' + str(len(rows1) + 7 + 4 + 2 + 12))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 13) + ':I' + str(len(rows1) + 7 + 4 + 2 + 14))

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10, len(rows1) + 7 + 4 + 2 + 10 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    fix_data6 = [
        [],
        [],
        ['四、线上交易分析', ],
        ['日期', '应收金额', '团购金额', '收单/买单金额', '团购占比',],
    ]
    for row in fix_data6:
        ws.append(row)

    ws.append(get_rows_sum4(rows1111))
    ws.append(get_rows_sum4(rows2222))
    fix_data7 = ['总结', '应收金额环比', '团购金额环比', '买单金额环比', ]
    ws.append(fix_data7)
    row_ratio = ['',
                 get_ratio(get_rows_sum4(rows1111)[1], get_rows_sum4(rows2222)[1]),
                 get_ratio(get_rows_sum4(rows1111)[2], get_rows_sum4(rows2222)[2]),
                 get_ratio(get_rows_sum4(rows1111)[3], get_rows_sum4(rows2222)[3]),
                 ]
    ws.append(row_ratio)
    summary4 = "总结：近" + str(len(rows11)) + "天（" + startdate1 + '到' + enddate1 + '）较上一轮' \
               + str(len(rows22)) + "天（" + startdate2 + '到' + enddate2 + '）' \
               + fix_data7[1] + get_change_text(row_ratio[1]) + ',' \
               + fix_data7[2] + get_change_text(row_ratio[2]) + ',' \
               + fix_data7[3] + get_change_text(row_ratio[3]) + ',' \
               + fix_data6[3][1] + str(Decimal(get_rows_sum4(rows1111)[1]).quantize(Decimal('0.00'))) + '.'

    ws.append([summary4])

    ws.cell(row=len(rows1) + 7 + 4 + 3 + 5 + 7 + 9, column=1).font = font_pass

    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 6) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10 + 6))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 7 + 5) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10 + 7 + 6))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 7 + 3) + ':A' + str(len(rows1) + 7 + 4 + 2 + 10 + 7 + 4))

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6, len(rows1) + 7 + 4 + 2 + 10 + 6 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6 + 4, len(rows1) + 7 + 4 + 2 + 10 + 6 + 3 + 4 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    fix_data8 = [
        [],
        [],
        ['五、店铺综合质量分（点评）', ],
        ['日期', '星级', '口味分', '环境分', '服务分', '好评增加', '中差评增加', '回复评价数'],
    ]
    for row in fix_data8:
        ws.append(row)

    for row in rows11111:
        ws.append(row)

    fix_data9 = ['', '总评论数', '好评率', '中差评率', '评价回复率', '累计好评增加', '累计中差评增加', '累计回复评价数', '日均评论数']
    ws.append(fix_data9)
    ws.append(get_rows_sum5(rows11111))
    # row_ratio = ['',
    #              get_ratio(get_rows_sum4(rows1111)[1], get_rows_sum4(rows2222)[1]),
    #              get_ratio(get_rows_sum4(rows1111)[2], get_rows_sum4(rows2222)[2]),
    #              get_ratio(get_rows_sum4(rows1111)[3], get_rows_sum4(rows2222)[3]),
    #              ]
    # ws.append(row_ratio)
    summary5 = "榜单分析：" + summary3[0] + "\n\n" + "1. 近" + str(len(rows11)) + "天（" + startdate1 + '到' + enddate1 + '）较上一轮' \
               + str(len(rows22)) + "天（" + startdate2 + '到' + enddate2 + '）' \
               + fix_data9[1] + str(get_rows_sum5(rows11111)[1]) + ',' \
               + fix_data9[2] + str(get_rows_sum5(rows11111)[2]) + ',' \
               + fix_data9[3] + str(get_rows_sum5(rows11111)[3]) + ',' \
               + fix_data9[4] + str(get_rows_sum5(rows11111)[4]) + ',' \
               + fix_data9[5] + str(get_rows_sum5(rows11111)[5]) + ',' \
               + fix_data9[6] + str(get_rows_sum5(rows11111)[6]) + ',' \
               + fix_data9[7] + str(get_rows_sum5(rows11111)[7]) + ',' \
               + fix_data8[3][1] + str((rows11111)[-1][1]) + '.'
    if get_rows_sum5(rows11111)[8] <= 2:
        summary5 = summary5 + "\n评论、包括好评数增加较少，望门店重视，可积极引导线下台卡活动。"

    summary5 = summary5 + "\n为提升ROS分，每天至少要增加两条好评，尽量避免中差评的出现，菜品质量和店内服务做要提升，评价中指出的问题可于次日班前开会提出并进行调整! "

    text1 = ""
    if float(get_ratio(get_rows_sum(rows1)[1], get_rows_sum(rows2)[1]).replace("%","")) > 0:
        text1 = "门店曝光有所提升，"
    else:
        text1 = "门店曝光相对减少，"

    summary5 = summary5 + "\n\n2. 近" + str(len(rows11)) + "天主要提到的问题：\n\n"

    summary5 = summary5 + "\n3. 近" + str(len(rows11)) + "天（" + startdate1 + '到' + enddate1 + '）较上一轮' \
               + str(len(rows22)) + "天（" + startdate2 + '到' + enddate2 + '）' \
               + text1 + fix_data7[1] + get_change_text(row_ratio[1]) + ',' \
               + fix_data7[2] + get_change_text(row_ratio[2]) + ',' \
               + fix_data7[3] + get_change_text(row_ratio[3]) + ',' \
               + fix_data6[3][1] + str(Decimal(get_rows_sum4(rows1111)[1]).quantize(Decimal('0.00'))) + '.'

    ws.append([summary5])

    ws.cell(row=len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 4 + len(rows11111), column=1).font = font_pass

    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 4 + len(rows11111)) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 4 + len(rows11111) + 15))

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6 + 9, len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 4 + len(rows11111), len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + 4 + len(rows11111) + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    fix_data10 = [
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        ['六、后续工作安排', ],
        ['''1、持续提升菜品质量，可定期研发新品，跟进门店新增评价；
            2、评价数据的监控，控制差评率；
            3、协助推广通投放监测提升ros质量分值'''],
    ]
    for row in fix_data10:
        ws.append(row)

    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 21) + ':I' + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 21))
    cell_merge.append("A" + str(len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 22) + ':I' + str(
        len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 22 + 10))

    Color_row_column.append(
        {
            'color': 'F2DCDB',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 21, len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 21 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    Color_row_column.append(
        {
            'color': 'C5D9F1',
            'cells': {
                'row': [len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 22,
                        len(rows1) + 7 + 4 + 2 + 10 + 6 + 9 + len(rows11111) + 22 + 1],
                'column': [1, 9 + 1]
            }
        }
    )

    row_sum = len(rows1) + 7 + 4 + 3 + 6 + 5 + 2 + 5 + 3 + len(rows1111) + 9 + 16

    for color_space in Color_row_column:
        fille = PatternFill('solid', fgColor=color_space['color'])
        for color_row in range(color_space['cells']['row'][0], color_space['cells']['row'][-1]):
            for color_column in range(color_space['cells']['column'][0], color_space['cells']['column'][-1]):
                ws.cell(row=color_row, column=color_column).fill = fille

    for cells in cell_merge:
        ws.merge_cells(cells)

    alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws['A1'].alignment = alignment_center

    ws.column_dimensions['A'].width = 30.0
    ws.row_dimensions[1].height = 40
    for i in range(2, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width = 20.0

    for i in ws["A1:I"+str(row_sum)]:
        for j in i:
            j.alignment = alignment_center



    print(sheetname + startdate1 + 'to' + enddate1 + 'with' + startdate2 + 'to' + enddate2 + '.xlsx' + '已保存')

    try:
        wb.save(sheetname + startdate1 + 'to' + enddate1 + 'with' + startdate2 + 'to' + enddate2 + '.xlsx')
    except:
        print('请关闭表格重新保存')


def get_data(filename, sheetname, startdate, enddate):
    wb_ck = openpyxl.load_workbook(filename + '.xlsx')
    ws_ck = wb_ck[sheetname]
    rows1 = []
    rows2 = []
    rows3 = []
    rows4 = []
    rows5 = []
    for row_ck in range(2, ws_ck.max_row + 1):
        if datetime.strptime(startdate, "%Y-%m-%d").date() <= datetime.strptime(ws_ck.cell(row_ck, 1).value,
                                                                                "%Y-%m-%d").date() <= datetime.strptime(
                enddate, "%Y-%m-%d").date():
            if int(ws_ck.cell(row_ck, 11).value.replace(",", "")) > 0:
                visit_rate = str(round(int(ws_ck.cell(row_ck, 5).value) / int(ws_ck.cell(row_ck, 11).value.replace(",", "")), 2) * 100) + "%"
            else:
                visit_rate = 0

            if int(ws_ck.cell(row_ck, 5).value) > 0:
                consume_rate = str(round(int(ws_ck.cell(row_ck, 4).value) / int(ws_ck.cell(row_ck, 5).value), 2) * 100) + "%"
            else:
                consume_rate = 0

            rows1.append([ws_ck.cell(row_ck, 1).value,
                         ws_ck.cell(row_ck, 11).value,
                         ws_ck.cell(row_ck, 5).value,
                         ws_ck.cell(row_ck, 4).value,
                         visit_rate,
                         consume_rate,
                         ws_ck.cell(row_ck, 2).value,
                         ws_ck.cell(row_ck, 10).value,
                         ws_ck.cell(row_ck, 13).value,
                         ])

            rows2.append([ws_ck.cell(row_ck, 1).value,
                          ws_ck.cell(row_ck, 14).value,
                          ws_ck.cell(row_ck, 15).value,
                          ws_ck.cell(row_ck, 16).value,

                          ws_ck.cell(row_ck, 21).value,

                          ])

            rows3.append([ws_ck.cell(row_ck, 1).value,
                          ws_ck.cell(row_ck, 38).value,
                          ws_ck.cell(row_ck, 39).value,
                          ])

            rows4.append([ws_ck.cell(row_ck, 1).value,
                          ws_ck.cell(row_ck, 17).value,
                          ws_ck.cell(row_ck, 19).value,
                          ws_ck.cell(row_ck, 20).value,

                          ])

            rows5.append([ws_ck.cell(row_ck, 1).value,
                          ws_ck.cell(row_ck, 27).value,

                          ws_ck.cell(row_ck, 28).value,
                          ws_ck.cell(row_ck, 29).value,
                          ws_ck.cell(row_ck, 30).value,

                          int(ws_ck.cell(row_ck, 24).value) - int(ws_ck.cell(row_ck, 25).value),
                          ws_ck.cell(row_ck, 25).value,
                          ws_ck.cell(row_ck, 24).value,

                          ])

    return rows1, rows2, rows3, rows4, rows5


def get_rows_sum2(rows):
    sum14 = 0
    sum15 = 0
    sum16 = 0
    sum21 = 0
    aver = 0
    pct = 0
    date_sum = rows[0][0] + '到' + rows[-1][0] + '合计'
    for row in rows:
        sum14 = sum14 + int(row[1])
        sum15 = sum15 + int(row[2])
        sum16 = sum16 + int(row[3])
        sum21 = sum21 + int(row[4])

    aver = sum16 / len(rows)

    if sum14 > 0:
        pct = str(round(sum15 / sum14, 2) * 100) + "%"
    else:
        pct = "0%"

    sums = [date_sum,
            sum14,
            sum15,
            sum16,
            aver,
            pct,
            sum21,
            ]

    return sums


def get_rows_sum4(rows):
    sum14 = 0
    sum15 = 0
    sum16 = 0
    pct = 0
    date_sum = rows[0][0] + '到' + rows[-1][0] + '合计'
    for row in rows:
        sum14 = sum14 + float(row[1])
        sum15 = sum15 + float(row[2])
        sum16 = sum16 + float(row[3])

    if sum14 > 0:
        pct = str(round(sum15 / sum14, 2) * 100) + "%"
    else:
        pct = "0%"

    sums = [date_sum,
            sum14,
            sum15,
            sum16,
            pct,
            ]

    return sums


def get_rows_sum5(rows):
    total = 0
    pct1 = 0
    pct2 = 0
    pct3 = 0
    sum14 = 0
    sum15 = 0
    sum16 = 0
    avg = 0

    date_sum = rows[0][0] + '到' + rows[-1][0] + '合计'
    for row in rows:
        sum14 = sum14 + int(row[5])
        sum15 = sum15 + int(row[6])
        sum16 = sum16 + int(row[7])

    total = sum16
    avg = total / len(rows)

    if total > 0:
        pct1 = str(round(sum14 / total, 2) * 100) + "%"
        pct2 = str(round(sum15 / total, 2) * 100) + "%"
        pct3 = str(round(sum16 / total, 2) * 100) + "%"
    else:
        pct1 = "0%"
        pct2 = "0%"
        pct3 = "0%"

    sums = [date_sum,
            total,
            pct1,
            pct2,
            pct3,
            sum14,
            sum15,
            sum16,
            avg,
            ]

    return sums


def get_ratio(value1, value2):
    if value2 > 0:
        ratio = Decimal((value1 - value2) / value2).quantize(Decimal('0.00')) * 100
    else:
        ratio = 0
    return str(ratio) + "%"


def get_change_text(text):
    if float(text.replace("%","")) >= 0:
        return_text = "上升" + str(abs(float(text.replace("%", "")), )) + "%"
    else:
        return_text = "下降" + str(abs(float(text.replace("%", "")), )) + "%"
    return return_text


def get_rows_sum(rows):
    sum2 = 0
    sum4 = 0
    sum5 = 0
    sum10 = 0
    sum11 = 0
    sum13 = 0
    date_sum = rows[0][0] + '到' + rows[-1][0] + '合计'
    for row in rows:
        sum2 = sum2 + int(row[1].replace(",", ""))
        sum4 = sum4 + int(row[2])
        sum5 = sum5 + int(row[3])
        sum10 = sum10 + float(row[6])
        sum11 = sum11 + int(row[7])
        # print(row[8])
        sum13 = sum13 + float(row[8])

    if sum2 > 0:
        visit_rate_sum = str(round(sum4 / sum2, 2) * 100) + "%"
    else:
        visit_rate_sum = 0

    if sum4 > 0:
        consume_rate_sum = str(round(sum5 / sum4, 2) * 100) + "%"
    else:
        consume_rate_sum = 0

    sums = [date_sum,
            sum2,
            sum4,
            sum5,
            visit_rate_sum,
            consume_rate_sum,
            sum10,
            sum11,
            sum13,
            ]

    return sums


def get_report(file):
    wb_ck = openpyxl.load_workbook(file)
    ws_ck = wb_ck['Sheet']

    for row_ck in range(1, ws_ck.max_row + 1):
        print(ws_ck.cell(row_ck, 1).value)
        file_name = ws_ck.cell(row_ck, 1).value
        sheet_name = ws_ck.cell(row_ck, 2).value
        start_date1 = ws_ck.cell(row_ck, 3).value
        end_date1 = ws_ck.cell(row_ck, 4).value
        start_date2 = ws_ck.cell(row_ck, 5).value
        end_date2 = ws_ck.cell(row_ck, 6).value
        get_template(file_name, sheet_name, start_date1, end_date1, start_date2, end_date2)


if __name__ == "__main__":
    get_report('report.xlsx')
